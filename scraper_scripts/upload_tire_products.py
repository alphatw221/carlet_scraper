import sys
# setting path
sys.path.append('../carlet_scraper')

import requests
import openpyxl
import re
import db
from PIL import Image

def login():
    account = 'celine'
    password = '8888good'
    url = 'https://service.gama.carlet.com.tw/api/anonymous/admin/login'

    response = requests.post(url,json={
        'account':account,
        'password':password
    })

    if response.status_code == 200:
        print("登入成功")
        # print("回應內容:", response.json())
        token = response.headers.get('X-Auth-Token')
        return token
    else:
        print("登入失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()

def is_valid_format(input_string):
    if type(input_string) != str:
        return False
    pattern = re.compile(r'^\d{3}/\d{2}/\d{2}$')
    return bool(pattern.match(input_string.strip()))


def create_price_table():
    brands=[{
            'name':'Bridgestone普利司通',
            'sheet_name':'Bridgestone',
            'series':[
                {
                    'name':'Ecopia節能王者',
                    'cost_column':1,
                    'price_column':2,
                    'price':{},
                    'cost':{},
                    'image_name':'bridgestone_ecopia.jpeg',
                    'description':\
                        '優越磨耗壽命，特殊化學結構減少滾動阻力，維持節能性更增加31%磨耗壽命。\n'\
                        +'四大特點：\n'\
                        +'1)更安靜-\n'\
                        +'  專為台灣消費者所設計，NH100 加深胎溝與'\
                        +'新花紋設計，讓輪胎與地面接觸的壓力平均化，'\
                        +'有效減少輪胎滾動時所產生的噪音，讓行駛過'\
                        +'程更安靜給予駕駛高品質靜肅性。\n'\
                        +'\n'\
                        +'2)更省油-\n'\
                        +'  全新「低滾動抗阻配方」，特殊胎面膠料結構'\
                        +'可讓降低分子磨擦與發熱，減少滾動阻抗，並'\
                        +'在保持抓地力同時讓輪胎的滾動距離增加，進'\
                        +'一步提昇節能表現。\n'\
                        +'\n'\
                        +'3)更耐磨-\n'\
                        +'  透過均衡的槽狀花紋和塊狀設計平衡胎塊的剛性，'\
                        +'讓輪胎接地壓力分佈平均避免偏磨耗的產生，藉'\
                        +'此來提升輪胎的耐磨耗特性。\n'\
                        +'\n'\
                        +'4)更安全-\n'\
                        +'  採用更深的主溝與胎肩深度，運用全新的直溝設'\
                        +'計，防止花紋塊變形，提高輪胎整體磨耗壽命。'
                    ,
                    'purchase_notes':\
                        '注意事項：\n'\
                        +'\n'\
                        +'使用期限\n'\
                        +'國家標準規定使用年限自2017/1/1開始實施：\n'\
                        +'1.超過製造日期6年之新輪胎不得安於汽車。\n'\
                        +'2.建議輪胎使用期限自製造日期起最長為10年\n'\
                        +'\n'\
                        +'更換輪胎\n'\
                        +'1.建議使用原廠出廠設定之尺寸、層級及乘載能力之輪胎。\n'\
                        +'2.組裝輪胎前，檢查鋼圈是否有異常損壞。\n'\
                        +'3.安裝輪胎時，必須於安全防爆籠中進行輪胎安裝和充氣，以防止輪胎爆炸。\n'\
                        +'4.為確保輪胎安裝正確，輪胎之鋼圈裝飾線和鋼圈邊緣之間的距離必須保持等距。\n'\
                        +'\n'\
                        +'輪胎充氣壓力\n'\
                        +'1.大多數輪胎故障是由不正確的胎壓所引起的，這嚴重影響了輪胎操控表現和輪胎壽命。\n'\
                        +'2.請於輪胎冷卻時調整胎壓。\n'\
                        +'3.避免在胎壓不足的情況下高速行駛。'
                },{
                    'name':'Turanza奢華舒適首選',
                    'cost_column':3,
                    'price_column':4,
                    'price':{},
                    'cost':{},
                    'image_name':'bridgestone_alenza.jpeg',
                    'description':
                        'Lexus LS/UX/ES 高級轎房車原廠配胎\n'\
                        +'1)有感的駕乘舒適性\n'\
                        +'  加強胎邊剛性，減少晃動帶來的不適感。同時透過花紋平均分配輪胎與路面的接觸壓力，抑制滾動噪音。\n'\
                        +'2)提升行駛安全性\n'\
                        +'  斜向花紋邊緣採用圓弧溝槽設計，防止胎塊變形或位移，提供輪胎更好的抓地力。\n'\
                        +'3)絕佳濕地性能-\n'\
                        +'  T005 煞車距離與其他品牌之平均煞車距離「減少3公尺」 。\n'\
                        +'  T005 半徑走行軌跡(R 42m) 「優於」其他品牌 (平均R 45,3m)。\n'\
                        +'4)歐盟TUV認證-  \n'\
                        +'  歐洲 TÜV 第三方認證 TURANZA T005 擁有最佳。'
                    ,
                    'purchase_notes':\
                        '注意事項：\n'\
                        +'\n'\
                        +'使用期限\n'\
                        +'國家標準規定使用年限自2017/1/1開始實施：\n'\
                        +'1.超過製造日期6年之新輪胎不得安於汽車。\n'\
                        +'2.建議輪胎使用期限自製造日期起最長為10年\n'\
                        +'\n'\
                        +'更換輪胎\n'\
                        +'1.建議使用原廠出廠設定之尺寸、層級及乘載能力之輪胎。\n'\
                        +'2.組裝輪胎前，檢查鋼圈是否有異常損壞。\n'\
                        +'3.安裝輪胎時，必須於安全防爆籠中進行輪胎安裝和充氣，以防止輪胎爆炸。\n'\
                        +'4.為確保輪胎安裝正確，輪胎之鋼圈裝飾線和鋼圈邊緣之間的距離必須保持等距。\n'\
                        +'\n'\
                        +'輪胎充氣壓力\n'\
                        +'1.大多數輪胎故障是由不正確的胎壓所引起的，這嚴重影響了輪胎操控表現和輪胎壽命。\n'\
                        +'2.請於輪胎冷卻時調整胎壓。\n'\
                        +'3.避免在胎壓不足的情況下高速行駛。'
                },{
                    'name':'Potenza高性能極致駕馭',
                    'cost_column':5,
                    'price_column':6,
                    'price':{},
                    'cost':{},
                    'image_name':'bridgestone_potenza.jpeg',
                    'description':\
                        '藍寶堅尼Huracán STO、瑪莎拉蒂MC20、BMW 8系列原廠指定配胎。\n'\
                        +'卓越乾地剎車-\n'\
                        +'採用了革新式橡膠混合技術，提升輪胎的乾濕地性能。\n'\
                        +'出色過彎性能&直行穩定性-\n'\
                        +'運動性胎面設計，提高穩定性與過彎性能，同時將滾動阻力最佳化。\n'\
                        +'出眾濕地過彎&操控-\n'\
                        +'胎面花紋排水設計提升角度彎道之轉彎性能。\n'\
                        +'六大特色-     \n'\
                        +'1)不對稱溝槽設計\n'\
                        +'  加強輪胎橫向剛性，減少轉彎時花紋塊變形及胎邊過度位移。\n'\
                        +'2)運動性能斷面設計\n'\
                        +'  超寬輪廓設計最大化接地面積，提升抓地力及乾地性能。\n'\
                        +'3)3D花紋設計\n'\
                        +'  提升花紋塊剛性進而強化剎車性能及減少胎面磨損。\n'\
                        +'4)絕佳的乾濕地操控\n'\
                        +'  透過排水溝槽加寬設計，在乾溼的路面條件下，提供更好的過彎和剎車性能。\n'\
                        +'5)精確的過彎性能\n'\
                        +'  胎面的寬直花紋塊設計，讓駕馭感受更靈敏、更精確，並提升駕駛樂趣。\n'\
                        +'6)剛性及穩定度提升\n'\
                        +'  透過肩部花紋塊連接，提升整體剛性，輪胎在高速時可以有更靈敏的操控反應及執行穩定性。\n'\
                    ,
                    'purchase_notes':\
                        '注意事項：\n'\
                        +'\n'\
                        +'使用期限\n'\
                        +'國家標準規定使用年限自2017/1/1開始實施：\n'\
                        +'1.超過製造日期6年之新輪胎不得安於汽車。\n'\
                        +'2.建議輪胎使用期限自製造日期起最長為10年\n'\
                        +'\n'\
                        +'更換輪胎\n'\
                        +'1.建議使用原廠出廠設定之尺寸、層級及乘載能力之輪胎。\n'\
                        +'2.組裝輪胎前，檢查鋼圈是否有異常損壞。\n'\
                        +'3.安裝輪胎時，必須於安全防爆籠中進行輪胎安裝和充氣，以防止輪胎爆炸。\n'\
                        +'4.為確保輪胎安裝正確，輪胎之鋼圈裝飾線和鋼圈邊緣之間的距離必須保持等距。\n'\
                        +'\n'\
                        +'輪胎充氣壓力\n'\
                        +'1.大多數輪胎故障是由不正確的胎壓所引起的，這嚴重影響了輪胎操控表現和輪胎壽命。\n'\
                        +'2.請於輪胎冷卻時調整胎壓。\n'\
                        +'3.避免在胎壓不足的情況下高速行駛。'
                },{
                    'name':'Alenza都會頂級SUV',
                    'cost_column':7,
                    'price_column':8,
                    'price':{},
                    'cost':{},
                    'image_name':'bridgestone_turanza.jpeg',
                    'description':\
                        '優越舒適性能-\n'\
                        +'新花紋平衡輪胎和道路接觸面，減少震動幅度。\n'\
                        +'出眾靜肅特性-\n'\
                        +'優化膠料降低輪胎噪音，提供駕駛優越的舒適體驗。\n'\
                        +'三大特色-\n'\
                        +'1)頂級舒適與靜肅性\n'\
                        +'  全新改良輪胎花紋及膠料，讓胎面能平衡與道路的接觸面，降低休旅車輪胎的振動進而減少噪音，不管駕駛於山區道路或山路，都能提供優越的舒適度與靜謐的乘坐體驗。\n'\
                        +'\n'\
                        +'2)提升濕地操控與煞車性能\n'\
                        +'  3D M型溝槽「花紋端圓弧化」用以抑止輪胎花紋塊的捲入，讓煞車時的接地壓力能平均分佈，也讓 ALENZA 輪胎的濕地煞車性能大幅進步。\n'\
                        +'\n'\
                        +'3)優越的磨耗壽\n'\
                        +'  ALENZA「多重圓形塊結構」優化直花紋邊緣處的圓度，當汽車轉向時能提高輪胎中心處的接地壓力，提高抓地力並且避免偏磨耗，高磨耗壽命＋低燃費性能，給予 SUV 休旅車最高規格輪胎選擇。'\
                        ,
                    'purchase_notes':\
                        '注意事項：\n'\
                        +'\n'\
                        +'使用期限\n'\
                        +'國家標準規定使用年限自2017/1/1開始實施：\n'\
                        +'1.超過製造日期6年之新輪胎不得安於汽車。\n'\
                        +'2.建議輪胎使用期限自製造日期起最長為10年\n'\
                        +'\n'\
                        +'更換輪胎\n'\
                        +'1.建議使用原廠出廠設定之尺寸、層級及乘載能力之輪胎。\n'\
                        +'2.組裝輪胎前，檢查鋼圈是否有異常損壞。\n'\
                        +'3.安裝輪胎時，必須於安全防爆籠中進行輪胎安裝和充氣，以防止輪胎爆炸。\n'\
                        +'4.為確保輪胎安裝正確，輪胎之鋼圈裝飾線和鋼圈邊緣之間的距離必須保持等距。\n'\
                        +'\n'\
                        +'輪胎充氣壓力\n'\
                        +'1.大多數輪胎故障是由不正確的胎壓所引起的，這嚴重影響了輪胎操控表現和輪胎壽命。\n'\
                        +'2.請於輪胎冷卻時調整胎壓。\n'\
                        +'3.避免在胎壓不足的情況下高速行駛。'
                },{
                    'name':'特殊規格系列',
                    'cost_column':9,
                    'price_column':10,
                    'price':{},
                    'cost':{},
                    'description':'特殊系列規格輪胎',
                    'image_name':'bridgestone_special.jpeg',
                    'purchase_notes':\
                        '注意事項：\n'\
                        +'\n'\
                        +'使用期限\n'\
                        +'國家標準規定使用年限自2017/1/1開始實施：\n'\
                        +'1.超過製造日期6年之新輪胎不得安於汽車。\n'\
                        +'2.建議輪胎使用期限自製造日期起最長為10年\n'\
                        +'\n'\
                        +'更換輪胎\n'\
                        +'1.建議使用原廠出廠設定之尺寸、層級及乘載能力之輪胎。\n'\
                        +'2.組裝輪胎前，檢查鋼圈是否有異常損壞。\n'\
                        +'3.安裝輪胎時，必須於安全防爆籠中進行輪胎安裝和充氣，以防止輪胎爆炸。\n'\
                        +'4.為確保輪胎安裝正確，輪胎之鋼圈裝飾線和鋼圈邊緣之間的距離必須保持等距。\n'\
                        +'\n'\
                        +'輪胎充氣壓力\n'\
                        +'1.大多數輪胎故障是由不正確的胎壓所引起的，這嚴重影響了輪胎操控表現和輪胎壽命。\n'\
                        +'2.請於輪胎冷卻時調整胎壓。\n'\
                        +'3.避免在胎壓不足的情況下高速行駛。'
                }
            ]
           
        },
        {
            'name':'Michelin米其林',
            'sheet_name':'Michelin',
            'series':[
                {
                    'name':'Primacy',
                    'cost_column':1,
                    'price_column':2,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''

                },{
                    'name':'Pilot Sport5',
                    'cost_column':3,
                    'price_column':4,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },{
                    'name':'Pilot Sport4s',
                    'cost_column':5,
                    'price_column':6,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },{
                    'name':'Pilot Sport4SUV',
                    'cost_column':7,
                    'price_column':8,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },
            ],
          
        },
        {
            'name':'Continental馬牌',
            'sheet_name':'Continental',
            'series':[{
                    'name':'ProContact7',
                    'cost_column':1,
                    'price_column':2,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },{
                    'name':'UltraContact7',
                    'cost_column':3,
                    'price_column':4,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },{
                    'name':'SportContact7',
                    'cost_column':5,
                    'price_column':6,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },{
                    'name':'ProContact7 SUV',
                    'cost_column':7,
                    'price_column':8,
                    'price':{},
                    'cost':{},
                    'description':'',
                    'purchase_notes':''
                },
            ]
            
        }
    ]

    excel_file_path = 'tires.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)

    for brand in brands:
        brand_name = brand.get('name')
        sheet_name = brand.get('sheet_name')
        series = brand.get('series')

        sheet = workbook[sheet_name]

        for serie in series:
            serie_name = serie.get('name')
            price_column = serie.get('price_column')
            cost_column = serie.get('cost_column')
            price_dict = serie.get('price')
            cost_dict = serie.get('cost')

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

                if is_valid_format(row[0].value):
                    if row[price_column].value:
                        price_dict[row[0].value.strip()] = float(row[price_column].value)
                    if row[cost_column].value:
                        cost_dict[row[0].value.strip()] = float(row[cost_column].value)

    workbook.close()
    return brands

def get_product(token, product_id):
    url = f'https://service.gama.carlet.com.tw/api/admin/store/parts/{product_id}'
    response = requests.get(url, headers={'Authorization':f'Bearer {token}'})
    if response.status_code == 200:
        data = response.json()
        entity = data.get('entity')
        return entity
    else:
        print("索取失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()

def check_product_exists(token, store, name):
    url = 'https://service.gama.carlet.com.tw/api/admin/store/partss'
    response = requests.get(url, params={'store':store, 'name':name}, headers={'Authorization':f'Bearer {token}'})
    if response.status_code == 200:
        data = response.json()
        entities = data.get('pager',{}).get('entities',[])
        if len(entities)==1:
            print(f"產品已存在 : {name}")
            return True, entities[0].get('id')
        elif not entities:
            return False, None
        else:
            raise Exception()

    else:
        print("查詢失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()

def create_product(token, store, name, hours, price, cost, description='其他說明', purchase_notes='購買說明'):
    url='https://service.gama.carlet.com.tw/api/admin/store/parts'
    data = {
        "store": store,
        "name": name,
        "hour": hours,
        "price": price,
        "cost": cost,
        "is_enabled": 1,
        "enabled_at": "2000-01-01 00:00:00",
        "disabled_at": "2099-12-31 23:59:59",
        "description": description,
        "purchase_notes": purchase_notes,
    }
    response = requests.post(url, json=data, headers={'Authorization':f'Bearer {token}'})

    if response.status_code == 200:
        print(f'新增產品成功 : {name}')
        data = response.json()
        return True, data.get('entity',{}).get('id')
    else:
        print("新增產品失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()

def add_product_image(token, product_id, image_name):
    url = f'https://service.gama.carlet.com.tw/api/admin/store/parts/{product_id}/image'

    with open(image_name, "rb") as file:
        print(file)
        response = requests.post(url, headers={'Authorization':f'Bearer {token}'}, files = {"file": file})
        if response.status_code == 200:
            print('新增圖片成功')
        else:
            print("新增圖片失敗")
            print("錯誤碼:", response.status_code)
            print("錯誤內容:", response.text)
            raise Exception()
    
def add_product_image_if_not_exists(token, product_id, image_name):

    product = get_product(token, product_id)
    if not product.get('photos',[]):
        add_product_image(token, product_id, image_name)
    else:
        print(f'商品{product_id}已存在照片')

def update_product_cateogry(token, product_id):
    url=f'https://service.gama.carlet.com.tw/api/admin/store/parts/{product_id}/categories'
    response = requests.put(url, json=[{
        'id':210,#輪胎
        'sid':999999,
    }], headers={'Authorization':f'Bearer {token}'})
    if response.status_code == 200:
        print('更新類別成功')
    else:
        print("更新類別失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()


def add_compatible_vehicle(token, product_id, vehicle_id):
    url = f'https://service.gama.carlet.com.tw/api/admin/store/parts/{product_id}/match/vehicles'
    response = requests.get(url, headers={'Authorization':f'Bearer {token}'})

    all, sizes, makes, models_include, models_exclude = False, [], [], [], []
    if response.status_code == 200:
        data = response.json()
        entity = data.get('entity',{})

        all = entity.get('all')
        sizes = entity.get('size',[])
        makes = [ make.get('name') for make in entity.get('make')]
        models_include = [ vehicle.get('id') for vehicle in entity.get('model_include')]
        models_exclude = [ vehicle.get('id') for vehicle in entity.get('model_exclude')]


    else:
        print("取得產品相容車款失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()
    
    if vehicle_id in models_include:
        print("新增產品相容車款成功")
        return
    
    models_include.append(vehicle_id)

    response = requests.put(url, headers={'Authorization':f'Bearer {token}'},
                             json={
                                    "all": all,
                                    "size": sizes,
                                    "make": makes,
                                    "model_include": models_include,
                                    "model_exclude": models_exclude
                             })
    
    if response.status_code == 200:
        print("新增產品相容車款成功")
    else:
        print("新增產品相容車款失敗")
        print("錯誤碼:", response.status_code)
        print("錯誤內容:", response.text)
        raise Exception()




def reformat(match):
    match = re.match(r'(\d{3}/\d{2})\s*([A-Z]+)\s*(\d{2})', match)
    formatted_spec = ''
    if match:
        groups = match.groups()
        formatted_spec = f"{groups[0]}/{groups[2]}"

    return formatted_spec

def extract_tire_sizes(input_string):

    pattern = re.compile(r'(\d{3}/\d{2}\s*[A-Z]+\s*\d{2})')
    matches = pattern.findall(input_string)
    return [reformat(match) for match in matches]





def get_tire_product_name(brand_name, serie_name, spec, variant_name):
    return f'{brand_name}{serie_name}-{spec}({variant_name})'


def upload_compactable_product(token:str, tire_price_data:list, store:str, vehicle:db.local_carlet.models.VehicleModel, specs:list, product_variants:list):
    
        for brand in tire_price_data:
            brand_name = brand.get('name')
            series = brand.get('series')

            for serie in series:
                serie_name = serie.get('name')
                price_dict = serie.get('price')
                cost_dict = serie.get('cost')
                description = serie.get('description')
                purchase_notes = serie.get('purchase_notes')
                
                for spec in specs:
                    
                    if spec not in price_dict or spec not in cost_dict:
                        print(f'{serie_name} 系列 沒有 {spec} 型號')
                        continue
                    
                    for product_variant in product_variants:
                        variant_name = product_variant.get('name')
                        factor = product_variant.get('factor')
                        hours = product_variant.get('hours')
                        tire_product_name = get_tire_product_name(brand_name, serie_name, spec, variant_name)
                        price = price_dict.get(spec) * factor
                        cost = cost_dict.get(spec) * factor

                        exists, product_id = check_product_exists(token, store, tire_product_name)

                        if not exists:
                            success, product_id = create_product(token, store, tire_product_name, hours, price, cost, description, purchase_notes,)
                            update_product_cateogry(token, product_id)

                        add_compatible_vehicle(token, product_id, vehicle.id)

def upload_all_tires():
    token = login()
    tire_price_data = create_price_table()
    store = 'Carlet輪胎館'

    product_variants = [
        {'name':'前輪2顆', 'factor':2, 'hours':'1'},
        {'name':'後輪2顆', 'factor':2, 'hours':'1'},
        {'name':'全車4顆', 'factor':4, 'hours':'2'}
    ]

    for brand in tire_price_data:
        brand_name = brand.get('name')
        series = brand.get('series')

        for serie in series:
            serie_name = serie.get('name')
            price_dict = serie.get('price')
            cost_dict = serie.get('cost')
            description = serie.get('description')
            purchase_notes = serie.get('purchase_notes')
            image_name = serie.get('image_name')

           
            for spec, price in price_dict.items():

                if spec not in cost_dict:
                    raise Exception()
                
                for product_variant in product_variants:
                    variant_name = product_variant.get('name')
                    factor = product_variant.get('factor')
                    hours = product_variant.get('hours')
                    tire_product_name = get_tire_product_name(brand_name, serie_name, spec, variant_name)
                    cost = cost_dict.get(spec)

                    price = price * factor
                    cost = cost * factor

                    exists, product_id = check_product_exists(token, store, tire_product_name)


                    if not exists:
                        success, product_id = create_product(token, store, tire_product_name, hours, price, cost, description, purchase_notes,)
                        update_product_cateogry(token, product_id)
                    
                    if image_name:
                        add_product_image_if_not_exists(token, product_id, image_name)



def main():
    token = login()
    tire_price_data = create_price_table()
    store = 'Carlet輪胎館'

    product_variants = [
        {'name':'前輪2顆', 'factor':2, 'hours':'1'},
        {'name':'後輪2顆', 'factor':2, 'hours':'1'},
        {'name':'全車4顆', 'factor':4, 'hours':'2'}
    ]


    with db.local_carlet.Session() as session:
        # vehicles= session.query(db.local_carlet.models.Vehicle).filter(db.local_carlet.models.Vehicle.make.in_(target_makes)).order_by(db.local_carlet.models.Vehicle.id.asc()).limit(50)
        vehicles= session.query(db.local_carlet.models.VehicleModel).filter(
            db.local_carlet.models.VehicleModel.auto_data_id!=None, 
            db.local_carlet.models.VehicleModel.auto_data_id!=-1,
            db.local_carlet.models.VehicleModel.mark==False,
            db.local_carlet.models.VehicleModel.mark3!=True,
            )\
        .yield_per(100)
        # .limit(5)
        
        

                              
    for vehicle in vehicles:
        print(vehicle.id)

        with db.auto_data.Session() as session:
            property = session.query(db.auto_data.car.Property).filter(db.auto_data.car.Property.car_id==vehicle.auto_data_id, db.auto_data.car.Property.name=='Tires size').first()
        
        if not property:
            with db.local_carlet.Session() as session:
                session.query(db.local_carlet.models.VehicleModel).filter_by(id=vehicle.id).update({'mark3': True})
                session.commit()
            continue

        if 'Rear wheel tires' in property.value:
            
            with db.local_carlet.Session() as session:
                session.query(db.local_carlet.models.VehicleModel).filter_by(id=vehicle.id).update({'mark2': True})
                session.commit()
            
            front_tires, rear_tires = property.value.split('Rear wheel tires')
            front_specs = extract_tire_sizes(front_tires)
            rear_specs = extract_tire_sizes(rear_tires)

            upload_compactable_product(
                token, 
                tire_price_data, 
                store, 
                vehicle, 
                front_specs, 
                [{'name':'前輪2顆', 'factor':2, 'hours':'1'}]
            )
            upload_compactable_product(
                token, 
                tire_price_data, 
                store, 
                vehicle, 
                rear_specs, 
                [{'name':'後輪2顆', 'factor':2, 'hours':'1'}]           
            )

        
        else:
            specs = extract_tire_sizes(property.value)
            upload_compactable_product(token, tire_price_data, store, vehicle, specs, product_variants)

        with db.local_carlet.Session() as session:
            session.query(db.local_carlet.models.VehicleModel).filter_by(id=vehicle.id).update({'mark': True})
            session.commit()

if __name__ == "__main__":
    # main()
    upload_all_tires()
