import sys
# setting path
sys.path.append('../carlet_scraper')

import openpyxl
import db
import re
from api import carlet_v1


def upload_compactable_product(token:str, tire_price_data:list, store:str, vehicle:db.local_carlet.models.VehicleModel, specs:list, product_variants:list):
    pass
        # for brand in tire_price_data:
        #     brand_name = brand.get('name')
        #     series = brand.get('series')

        #     for serie in series:
        #         serie_name = serie.get('name')
        #         price_dict = serie.get('price')
        #         cost_dict = serie.get('cost')
        #         description = serie.get('description')
        #         purchase_notes = serie.get('purchase_notes')
                
        #         for spec in specs:
                    
        #             if spec not in price_dict or spec not in cost_dict:
        #                 print(f'{serie_name} 系列 沒有 {spec} 型號')
        #                 continue
                    
        #             for product_variant in product_variants:
        #                 variant_name = product_variant.get('name')
        #                 factor = product_variant.get('factor')
        #                 hours = product_variant.get('hours')
        #                 tire_product_name = get_tire_product_name(brand_name, serie_name, spec, variant_name)
        #                 price = price_dict.get(spec) * factor
        #                 cost = cost_dict.get(spec) * factor

        #                 exists, product_id = check_product_exists(token, store, tire_product_name)

        #                 if not exists:
        #                     success, product_entity = create_product(token, store, tire_product_name, hours, price, cost, description, purchase_notes,)
        #                     update_product_cateogry(token, product_entity.get('id'))

        #                 add_compatible_vehicle(token, product_id, vehicle.id)








def create_oil_quotation()->dict:
    excel_file_path = 'spread_sheets/The Reset Auto小保養(機油).xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)

    oil_brands = {}
    for worksheet in workbook.worksheets:
        oil_products = {}
        oil_brands[worksheet.title] = oil_products
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=0, max_col=worksheet.max_column):
            oil_vis, price, cost, description, purchase_notes= row[0].value, row[1].value, row[2].value, row[3].value, row[4].value
            oil_products[oil_vis] = {'cost':cost, 'price':price, 'description':description, 'purchase_notes':purchase_notes}
    
    return oil_brands

def create_vehicle_quotation()->dict:
    excel_file_path = 'spread_sheets/The Reset Auto小保養(引擎).xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)


    makes = {}
    for worksheet in workbook.worksheets:
        vehicles = []
        makes[worksheet.title] = vehicles

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=0, max_col=worksheet.max_column):
            year, model, submodel, engine, exchange_volume, oil_brand, oil_vis, oil_filter_price, oil_filter_cost = row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value
            
            if not oil_brand or not oil_vis or not oil_filter_price or not oil_filter_cost:
                continue

            vehicles.append({
                'year':year,
                'model':model,
                'submodel':submodel,
                'engine':engine,
                'exchange_volume':exchange_volume,
                'oil_brand':oil_brand,
                'oil_vis':oil_vis,
                'oil_filter_price':oil_filter_price,
                'oil_filter_cost':oil_filter_cost,
            })

    return makes 



def search_match_vehicles_with_engine(make, engine):

    with db.local_carlet.Session() as session:
        vehicles = session.query(db.local_carlet.models.VehicleModel)\
            .join(db.local_carlet.models.VehicleMake)\
            .filter(db.local_carlet.models.VehicleMake.name==make, 
                    db.local_carlet.models.VehicleModel.engine==engine)
    return vehicles

def search_match_vehicles(make, year, model, submodel):

    with db.local_carlet.Session() as session:
        vehicles = session.query(db.local_carlet.models.VehicleModel)\
            .join(db.local_carlet.models.VehicleMake)\
            .filter(db.local_carlet.models.VehicleMake.name==make, 
                    db.local_carlet.models.VehicleModel.year==year,
                    db.local_carlet.models.VehicleModel.name==model,
                    db.local_carlet.models.VehicleModel.name_variant==submodel
                    )
    return vehicles


def _extract_exchange_volumn_liter(string):
    match = re.search(r'(\d+(\.\d+)?) l\n', string)

    if match:
        result = match.group(1)
        return float(result)
    else:
        print("未找到數字")
        raise Exception()
    
def get_exchange_volume_with_engine(make, engine):
    with db.local_carlet.Session() as session:
        vehicle = session.query(db.local_carlet.models.VehicleModel)\
            .join(db.local_carlet.models.VehicleMake)\
            .filter(db.local_carlet.models.VehicleMake.name==make, 
                    db.local_carlet.models.VehicleModel.engine==engine,
                    db.local_carlet.models.VehicleModel.auto_data_id!=None,
                    db.local_carlet.models.VehicleModel.auto_data_id>0
                    ).first()

    if not vehicle:
        return False, None
    
    with db.auto_data.Session() as session:
        property = session.query(db.auto_data.car.Property).filter(db.auto_data.car.Property.car_id==vehicle.auto_data_id, db.auto_data.car.Property.name=='Engine oil capacity').first()

    if not property:
        return False, None
    
    return True, _extract_exchange_volumn_liter(property.value)

def get_name(oil_brand, oil_vis, engine):
    return f'小保養{oil_brand}{oil_vis}（{engine}引擎）'

def main():
    token = carlet_v1.login()
    oil_quotation:dict = create_oil_quotation()
    vehicle_quotation:dict = create_vehicle_quotation()
    store = 'Carlet保養館'

    for make, vehicles in vehicle_quotation.items():

        for vehicle in vehicles:
            year = vehicle.get('year')
            model = vehicle.get('model')
            submodel = vehicle.get('submodel')
            engine = vehicle.get('engine')
            exchange_volume = vehicle.get('exchange_volume')
            oil_brand = vehicle.get('oil_brand')
            oil_vis = vehicle.get('oil_vis')
            oil_filter_price = vehicle.get('oil_filter_price')
            oil_filter_cost = vehicle.get('oil_filter_cost')

            if not engine:
                continue
            else:
                match_vehicles = search_match_vehicles_with_engine(make, engine)

            if not exchange_volume:
                success, exchange_volume = get_exchange_volume_with_engine(make, engine)
                if not success:
                    raise Exception()

            name = get_name(oil_brand, oil_vis, engine)

            exists, product_id = carlet_v1.check_product_exists(token, store, name)

            if not exists:

                oil = oil_quotation.get(oil_brand,{}).get(oil_vis)
                
                if not oil:
                    raise Exception()

                price = float(oil.get('price')) * float(exchange_volume) + float(oil_filter_price)
                cost = float(oil.get('cost')) * float(exchange_volume) + float(oil_filter_cost)
                description = oil.get('description')
                purchase_notes = oil.get('purchase_notes')
                
                success, product_entity = carlet_v1.create_product(token, store, name, '4', price, cost, description, purchase_notes)
                product_id = product_entity.get('id')
                carlet_v1.update_product_cateogry(token, product_id, 208)
                carlet_v1.add_product_image(token, product_id, 'images/客樂科技_保養.png')

            

            vehicle_ids = [match_vehicle.id for match_vehicle in match_vehicles]
            carlet_v1.add_compatible_vehicles(token, product_id, vehicle_ids)


          


if __name__ == "__main__":
    main()
